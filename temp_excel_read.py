import sys
try:
    import openpyxl
except Exception as e:
    print('OPENPYXL_MISSING', e)
    sys.exit(1)

path = r'C:\Users\MY LENOVO\Downloads\Thunderbit_be0e3c_20260604_032849.xlsx'
try:
    workbook = openpyxl.load_workbook(path, data_only=True)
    print('SHEETS:', workbook.sheetnames)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        headers = rows[0] if rows else []
        print('---', sheet_name, '---')
        print(','.join(str(h) for h in headers))
        for row in rows[1:11]:
            print(','.join('' if v is None else str(v) for v in row))
except Exception as e:
    print('ERROR', type(e).__name__, e)
